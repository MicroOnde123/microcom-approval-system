from io import BytesIO

from django.contrib.auth import get_user_model
from django.core import mail
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook
from urllib.parse import quote

from accounts.models import Department
from inventory.models import Material, MaterialCategory
from requests_app.models import (
    Request,
    RequestApproval,
    RequestAuditLog,
    RequestMaterialItem,
    RequestType,
)
from requests_app.services import submit_request
from workflows.models import ApprovalWorkflow, ApprovalWorkflowStep


TWO_COPY_WARNING = (
    "This request has too many material items for two-copy printing. "
    "Please print one copy."
)


class AlternateApproverWorkflowTests(TestCase):
    def setUp(self):
        self.department = Department.objects.create(
            name="Technical",
            code="TECH",
        )

        User = get_user_model()
        self.submitter = User.objects.create_user(
            username="submitter",
            password="pass12345",
            email="submitter@example.com",
            full_name="Submitter User",
            department=self.department,
        )
        self.primary = User.objects.create_user(
            username="primary",
            password="pass12345",
            email="primary@example.com",
            full_name="Primary Approver",
        )
        self.alternate = User.objects.create_user(
            username="alternate",
            password="pass12345",
            email="alternate@example.com",
            full_name="Alternate Approver",
        )

        self.request_type = RequestType.objects.create(
            name="Any Display Name",
            code="ANY",
            is_active=True,
        )
        self.workflow = ApprovalWorkflow.objects.create(
            name="Default approval",
            request_type=self.request_type,
            department=self.department,
            is_active=True,
        )
        self.workflow_step = ApprovalWorkflowStep.objects.create(
            workflow=self.workflow,
            step_order=1,
            approver_user=self.primary,
            alternate_approver_user=self.alternate,
        )

    def make_submitted_request(self):
        request_obj = Request.objects.create(
            request_number=f"REQ-{Request.objects.count() + 1}",
            request_type=self.request_type,
            submitted_by=self.submitter,
            department=self.department,
            description="Needs approval",
            status="PENDING",
        )
        submit_request(request_obj)
        return request_obj

    def get_approval(self, request_obj):
        return RequestApproval.objects.get(request=request_obj)

    def act_on_approval(self, user, approval, action):
        self.client.force_login(user)
        return self.client.post(
            reverse("approval_detail", args=[approval.id]),
            data={
                "action": action,
                "comment": f"{action} comment",
            },
            follow=True,
            HTTP_HOST="127.0.0.1",
        )

    def pending_count_for(self, user):
        self.client.force_login(user)
        response = self.client.get(
            reverse("notification_count"),
            HTTP_HOST="127.0.0.1",
        )
        return response.json()["pending"]

    def assert_action_result(self, request_obj, actor, action, expected_status):
        approval = self.get_approval(request_obj)
        response = self.act_on_approval(actor, approval, action)

        self.assertEqual(response.status_code, 200)

        approval.refresh_from_db()
        request_obj.refresh_from_db()

        self.assertEqual(approval.status, expected_status)
        self.assertEqual(approval.acted_by, actor)
        self.assertIsNotNone(approval.acted_at)
        self.assertEqual(request_obj.status, expected_status)

    def test_primary_approver_can_approve(self):
        request_obj = self.make_submitted_request()

        self.assert_action_result(request_obj, self.primary, "APPROVE", "APPROVED")

    def test_alternate_approver_can_approve(self):
        request_obj = self.make_submitted_request()

        self.assert_action_result(request_obj, self.alternate, "APPROVE", "APPROVED")

    def test_primary_approver_can_reject(self):
        request_obj = self.make_submitted_request()

        self.assert_action_result(request_obj, self.primary, "REJECT", "REJECTED")

    def test_alternate_approver_can_reject(self):
        request_obj = self.make_submitted_request()

        self.assert_action_result(request_obj, self.alternate, "REJECT", "REJECTED")

    def test_primary_approver_can_return(self):
        request_obj = self.make_submitted_request()

        self.assert_action_result(request_obj, self.primary, "RETURN", "RETURNED")

    def test_alternate_approver_can_return(self):
        request_obj = self.make_submitted_request()

        self.assert_action_result(request_obj, self.alternate, "RETURN", "RETURNED")

    def test_once_approved_by_primary_alternate_cannot_approve(self):
        request_obj = self.make_submitted_request()
        approval = self.get_approval(request_obj)

        self.act_on_approval(self.primary, approval, "APPROVE")
        response = self.act_on_approval(self.alternate, approval, "APPROVE")

        approval.refresh_from_db()
        self.assertEqual(approval.status, "APPROVED")
        self.assertEqual(approval.acted_by, self.primary)
        self.assertContains(
            response,
            "This step has already been acted upon.",
            status_code=200,
        )

    def test_once_approved_by_alternate_primary_cannot_approve(self):
        request_obj = self.make_submitted_request()
        approval = self.get_approval(request_obj)

        self.act_on_approval(self.alternate, approval, "APPROVE")
        response = self.act_on_approval(self.primary, approval, "APPROVE")

        approval.refresh_from_db()
        self.assertEqual(approval.status, "APPROVED")
        self.assertEqual(approval.acted_by, self.alternate)
        self.assertContains(
            response,
            "This step has already been acted upon.",
            status_code=200,
        )

    def test_pending_approvals_disappear_for_other_user_after_action(self):
        request_obj = self.make_submitted_request()
        approval = self.get_approval(request_obj)

        self.assertEqual(self.pending_count_for(self.primary), 1)
        self.assertEqual(self.pending_count_for(self.alternate), 1)

        self.act_on_approval(self.primary, approval, "APPROVE")

        self.assertEqual(self.pending_count_for(self.primary), 0)
        self.assertEqual(self.pending_count_for(self.alternate), 0)

    def test_notification_counts_work_for_primary_and_alternate(self):
        request_obj = self.make_submitted_request()
        approval = self.get_approval(request_obj)

        self.assertEqual(self.pending_count_for(self.primary), 1)
        self.assertEqual(self.pending_count_for(self.alternate), 1)

        self.act_on_approval(self.alternate, approval, "REJECT")

        self.assertEqual(self.pending_count_for(self.primary), 0)
        self.assertEqual(self.pending_count_for(self.alternate), 0)

    def test_submit_request_sends_email_to_primary_and_alternate(self):
        self.make_submitted_request()

        self.assertEqual(len(mail.outbox), 1)
        self.assertEqual(
            set(mail.outbox[0].to),
            {"primary@example.com", "alternate@example.com"},
        )

    def test_primary_approver_name_is_displayed_as_actor(self):
        request_obj = self.make_submitted_request()
        approval = self.get_approval(request_obj)

        self.act_on_approval(self.primary, approval, "APPROVE")

        self.client.force_login(self.primary)
        response = self.client.get(
            reverse("request_detail", args=[request_obj.id]),
            HTTP_HOST="127.0.0.1",
        )

        self.assertContains(response, "Primary Approver")

    def test_alternate_approver_name_is_displayed_as_actor(self):
        request_obj = self.make_submitted_request()
        approval = self.get_approval(request_obj)

        self.act_on_approval(self.alternate, approval, "APPROVE")

        self.client.force_login(self.alternate)
        response = self.client.get(
            reverse("request_detail", args=[request_obj.id]),
            HTTP_HOST="127.0.0.1",
        )

        self.assertContains(response, "Alternate Approver")
        self.assertNotContains(response, "Primary Approver")

    def test_request_detail_shows_review_button_to_primary_approver(self):
        request_obj = self.make_submitted_request()
        approval = self.get_approval(request_obj)
        detail_url = (
            f"{reverse('request_detail', args=[request_obj.id])}"
            "?next=%2Fapprovals%2Fpending%2F%3Fstatus%3DPENDING"
        )

        self.client.force_login(self.primary)
        response = self.client.get(detail_url, HTTP_HOST="127.0.0.1")

        expected_url = (
            f"{reverse('approval_detail', args=[approval.id])}"
            f"?next={quote(detail_url, safe='/')}"
        )
        self.assertContains(response, "Review / Approve Request")
        self.assertContains(response, expected_url)
        self.assertEqual(response.context["active_approval_for_user"], approval)

    def test_request_detail_shows_review_button_to_alternate_approver(self):
        request_obj = self.make_submitted_request()
        approval = self.get_approval(request_obj)

        self.client.force_login(self.alternate)
        response = self.client.get(
            reverse("request_detail", args=[request_obj.id]),
            HTTP_HOST="127.0.0.1",
        )

        self.assertContains(response, "Review / Approve Request")
        self.assertEqual(response.context["active_approval_for_user"], approval)

    def test_request_detail_hides_review_button_from_requester(self):
        request_obj = self.make_submitted_request()

        self.client.force_login(self.submitter)
        response = self.client.get(
            reverse("request_detail", args=[request_obj.id]),
            HTTP_HOST="127.0.0.1",
        )

        self.assertNotContains(response, "Review / Approve Request")
        self.assertIsNone(response.context["active_approval_for_user"])

    def test_request_detail_hides_review_button_from_later_step_approver(self):
        later_approver = get_user_model().objects.create_user(
            username="later-approver",
            password="pass12345",
            email="later@example.com",
            full_name="Later Approver",
        )
        ApprovalWorkflowStep.objects.create(
            workflow=self.workflow,
            step_order=2,
            approver_user=later_approver,
        )
        request_obj = self.make_submitted_request()

        self.client.force_login(later_approver)
        response = self.client.get(
            reverse("request_detail", args=[request_obj.id]),
            HTTP_HOST="127.0.0.1",
        )

        self.assertNotContains(response, "Review / Approve Request")
        self.assertIsNone(response.context["active_approval_for_user"])

    def test_request_detail_hides_review_button_from_unassigned_admin(self):
        request_obj = self.make_submitted_request()
        admin = get_user_model().objects.create_superuser(
            username="admin",
            password="pass12345",
            email="admin@example.com",
            full_name="Admin User",
        )

        self.client.force_login(admin)
        response = self.client.get(
            reverse("request_detail", args=[request_obj.id]),
            HTTP_HOST="127.0.0.1",
        )

        self.assertNotContains(response, "Review / Approve Request")
        self.assertIsNone(response.context["active_approval_for_user"])

    def test_request_detail_hides_review_button_from_unassigned_stock_manager(self):
        request_obj = self.make_submitted_request()
        stock_manager = get_user_model().objects.create_user(
            username="stock-manager",
            password="pass12345",
            email="stock@example.com",
            full_name="Stock Manager",
            can_manage_stock=True,
        )
        category = MaterialCategory.objects.create(name="Office", code="OFFICE")
        material = Material.objects.create(
            category=category,
            name="Paper",
            code="PAPER",
        )
        RequestMaterialItem.objects.create(
            request=request_obj,
            material=material,
            quantity=1,
        )

        self.client.force_login(stock_manager)
        response = self.client.get(
            reverse("request_detail", args=[request_obj.id]),
            HTTP_HOST="127.0.0.1",
        )

        self.assertNotContains(response, "Review / Approve Request")
        self.assertIsNone(response.context["active_approval_for_user"])

    def test_request_detail_hides_review_button_after_approval(self):
        request_obj = self.make_submitted_request()
        approval = self.get_approval(request_obj)
        self.act_on_approval(self.primary, approval, "APPROVE")

        self.client.force_login(self.primary)
        response = self.client.get(
            reverse("request_detail", args=[request_obj.id]),
            HTTP_HOST="127.0.0.1",
        )

        self.assertNotContains(response, "Review / Approve Request")
        self.assertIsNone(response.context["active_approval_for_user"])

    def test_historical_approval_without_actor_falls_back_to_assigned_approver(self):
        request_obj = self.make_submitted_request()
        approval = self.get_approval(request_obj)
        approval.status = "APPROVED"
        approval.acted_by = None
        approval.acted_at = timezone.now()
        approval.save()
        request_obj.status = "APPROVED"
        request_obj.current_step_order = None
        request_obj.finalized_at = timezone.now()
        request_obj.save()

        self.client.force_login(self.primary)
        response = self.client.get(
            reverse("request_detail", args=[request_obj.id]),
            HTTP_HOST="127.0.0.1",
        )

        self.assertContains(response, "Primary Approver")

    def test_historical_approval_without_actor_appears_in_assigned_approver_history(self):
        request_obj = self.make_submitted_request()
        approval = self.get_approval(request_obj)
        approval.status = "APPROVED"
        approval.acted_by = None
        approval.acted_at = timezone.now()
        approval.save()
        request_obj.status = "APPROVED"
        request_obj.current_step_order = None
        request_obj.finalized_at = timezone.now()
        request_obj.save()

        self.client.force_login(self.primary)
        response = self.client.get(
            reverse("approval_history"),
            HTTP_HOST="127.0.0.1",
        )

        self.assertContains(response, request_obj.request_number, count=1)

    def test_pending_approvals_shows_request_once_when_user_is_primary_and_alternate(self):
        self.workflow_step.alternate_approver_user = self.primary
        self.workflow_step.save()
        request_obj = self.make_submitted_request()

        self.client.force_login(self.primary)
        response = self.client.get(
            reverse("pending_approvals"),
            HTTP_HOST="127.0.0.1",
        )

        self.assertContains(response, request_obj.request_number, count=1)
        self.assertEqual(self.pending_count_for(self.primary), 1)

    def test_approval_history_shows_request_once_for_multi_step_workflow(self):
        ApprovalWorkflowStep.objects.create(
            workflow=self.workflow,
            step_order=2,
            approver_user=self.primary,
            alternate_approver_user=self.alternate,
        )
        request_obj = self.make_submitted_request()

        first_approval = RequestApproval.objects.get(
            request=request_obj,
            step_order=1,
        )
        self.act_on_approval(self.primary, first_approval, "APPROVE")

        second_approval = RequestApproval.objects.get(
            request=request_obj,
            step_order=2,
        )
        self.act_on_approval(self.primary, second_approval, "APPROVE")

        self.client.force_login(self.primary)
        response = self.client.get(
            reverse("approval_history"),
            HTTP_HOST="127.0.0.1",
        )

        self.assertContains(response, request_obj.request_number, count=1)


class MaterialPrintCopyLimitTests(TestCase):
    def setUp(self):
        self.department = Department.objects.create(
            name="Technical",
            code="TECH",
        )

        User = get_user_model()
        self.submitter = User.objects.create_user(
            username="material_submitter",
            password="pass12345",
            email="material-submit@example.com",
            full_name="Material Submitter",
            department=self.department,
        )
        self.stock_user = User.objects.create_user(
            username="stock_user",
            password="pass12345",
            email="stock@example.com",
            full_name="Stock User",
            can_manage_stock=True,
        )

        self.request_type = RequestType.objects.create(
            name="Material Request",
            code="MAT",
            is_active=True,
            requires_materials=True,
        )
        self.category = MaterialCategory.objects.create(
            name="General",
            code="GEN",
        )

    def make_material_request(self, item_count, number_suffix):
        request_obj = Request.objects.create(
            request_number=f"MAT-{number_suffix}",
            request_type=self.request_type,
            submitted_by=self.submitter,
            department=self.department,
            description="For installation",
            date_needed=timezone.now().date(),
            status="APPROVED",
            finalized_at=timezone.now(),
        )

        for index in range(item_count):
            material = Material.objects.create(
                category=self.category,
                name=f"Material {number_suffix}-{index}",
                code=f"MAT-{number_suffix}-{index}",
                unit="pcs",
                stock_quantity=10,
            )
            RequestMaterialItem.objects.create(
                request=request_obj,
                material=material,
                quantity=1,
            )

        return request_obj

    def assert_material_slip_count(self, response, count):
        content = response.content.decode()
        self.assertEqual(content.count("Material Exit Slip"), count)

    def export_workbook(self, language):
        self.client.force_login(self.stock_user)

        response = self.client.get(
            reverse("export_material_report_excel"),
            HTTP_ACCEPT_LANGUAGE=language,
            HTTP_HOST="127.0.0.1",
        )

        workbook = load_workbook(BytesIO(response.content))
        return response, workbook.active

    def test_material_reports_excel_and_print_controls_are_localized(self):
        self.client.force_login(self.stock_user)

        english_response = self.client.get(
            reverse("material_reports"),
            HTTP_ACCEPT_LANGUAGE="en",
            HTTP_HOST="127.0.0.1",
        )
        self.assertContains(english_response, "Export Excel")
        self.assertContains(english_response, "Print selected: 1 copy")
        self.assertContains(english_response, "Print selected: 2 copies")

        french_response = self.client.get(
            reverse("material_reports"),
            HTTP_ACCEPT_LANGUAGE="fr",
            HTTP_HOST="127.0.0.1",
        )
        self.assertContains(french_response, "Exporter Excel")
        self.assertContains(french_response, "Imprimer la sélection : 1 exemplaire")
        self.assertContains(french_response, "Imprimer la sélection : 2 exemplaires")

    def test_excel_export_uses_english_filename_and_labels(self):
        response, sheet = self.export_workbook("en")

        self.assertEqual(
            response["Content-Disposition"],
            'attachment; filename="material_report.xlsx"',
        )
        self.assertEqual(sheet.title, "Material Report")
        self.assertEqual(sheet["A1"].value, "Microcom Material Report")
        self.assertEqual(sheet["A2"].value, "Generated At")
        self.assertEqual(
            [sheet.cell(row=4, column=column).value for column in range(1, 15)],
            [
                "Request Number",
                "Requester",
                "Department",
                "Date Needed",
                "Approved Date",
                "Material",
                "Material Code",
                "Category",
                "Quantity",
                "Unit",
                "Available Stock",
                "Description",
                "Material Issue Note",
                "Approvers",
            ],
        )

    def test_excel_export_uses_french_filename_and_labels(self):
        response, sheet = self.export_workbook("fr")

        self.assertEqual(
            response["Content-Disposition"],
            'attachment; filename="rapport_materiel.xlsx"',
        )
        self.assertEqual(sheet.title, "Rapport de matériel")
        self.assertEqual(sheet["A1"].value, "Rapport de matériel Microcom")
        self.assertEqual(sheet["A2"].value, "Généré le")
        self.assertEqual(
            [sheet.cell(row=4, column=column).value for column in range(1, 15)],
            [
                "Numéro de demande",
                "Demandeur",
                "Département",
                "Date requise",
                "Date d’approbation",
                "Matériel",
                "Code matériel",
                "Catégorie",
                "Quantité",
                "Unité",
                "Stock disponible",
                "Description",
                "Note de sortie matériel",
                "Approbateurs",
            ],
        )

    def test_single_material_print_forces_one_copy_when_more_than_six_items(self):
        request_obj = self.make_material_request(item_count=7, number_suffix="007")
        self.client.force_login(self.submitter)

        response = self.client.get(
            f"{reverse('approved_document', args=[request_obj.id])}?copies=2",
            HTTP_HOST="127.0.0.1",
        )

        self.assertContains(response, TWO_COPY_WARNING)
        self.assertNotContains(response, 'class="print-sheet two-copies"')
        self.assert_material_slip_count(response, 1)

    def test_single_material_print_allows_two_copies_at_six_items(self):
        request_obj = self.make_material_request(item_count=6, number_suffix="006")
        self.client.force_login(self.submitter)

        response = self.client.get(
            f"{reverse('approved_document', args=[request_obj.id])}?copies=2",
            HTTP_HOST="127.0.0.1",
        )

        self.assertNotContains(response, TWO_COPY_WARNING)
        self.assertContains(response, 'class="print-sheet two-copies"')
        self.assert_material_slip_count(response, 2)

    def test_bulk_material_print_forces_one_copy_only_for_oversized_requests(self):
        small_request = self.make_material_request(item_count=2, number_suffix="002")
        large_request = self.make_material_request(item_count=7, number_suffix="107")
        self.client.force_login(self.stock_user)

        response = self.client.post(
            f"{reverse('bulk_print_material_documents')}?copies=2",
            data={
                "selected_requests": [small_request.id, large_request.id],
            },
            HTTP_HOST="127.0.0.1",
        )

        self.assertContains(response, TWO_COPY_WARNING)
        self.assert_material_slip_count(response, 3)
        self.assertContains(response, 'class="print-sheet two-copies"', count=1)


class MaterialIssueNoteTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.department = Department.objects.create(name="Field Services", code="FIELD")
        User = get_user_model()
        cls.requester = User.objects.create_user(
            username="note-requester",
            password="pass12345",
            email="note-requester@example.com",
            full_name="Note Requester",
            department=cls.department,
        )
        cls.approver = User.objects.create_user(
            username="note-approver",
            password="pass12345",
            email="note-approver@example.com",
            full_name="Note Approver",
        )
        cls.stock_manager = User.objects.create_user(
            username="note-stock-manager",
            password="pass12345",
            email="note-stock@example.com",
            full_name="Note Stock Manager",
            can_manage_stock=True,
        )
        cls.regular_user = User.objects.create_user(
            username="note-regular",
            password="pass12345",
            email="note-regular@example.com",
            full_name="Note Regular User",
        )
        cls.superuser = User.objects.create_superuser(
            username="note-superuser",
            password="pass12345",
            email="note-admin@example.com",
            full_name="Note Superuser",
        )
        cls.material_type = RequestType.objects.create(
            name="Material Request",
            code="NOTE-MATERIAL",
            requires_materials=True,
        )
        cls.non_material_type = RequestType.objects.create(
            name="Service Request",
            code="NOTE-SERVICE",
            requires_materials=False,
        )
        cls.category = MaterialCategory.objects.create(name="Network", code="NOTE-NET")
        cls.material = Material.objects.create(
            category=cls.category,
            name="Router",
            code="NOTE-ROUTER",
            unit="pcs",
            stock_quantity=5,
        )
        cls.request_obj = Request.objects.create(
            request_number="NOTE-001",
            request_type=cls.material_type,
            submitted_by=cls.requester,
            department=cls.department,
            description="Install customer router",
            status="APPROVED",
            finalized_at=timezone.now(),
            material_issue_note="SN-OLD",
        )
        RequestMaterialItem.objects.create(
            request=cls.request_obj,
            material=cls.material,
            quantity=1,
        )
        workflow = ApprovalWorkflow.objects.create(
            name="Note approval",
            request_type=cls.material_type,
            department=cls.department,
        )
        workflow_step = ApprovalWorkflowStep.objects.create(
            workflow=workflow,
            step_order=1,
            approver_user=cls.approver,
        )
        RequestApproval.objects.create(
            request=cls.request_obj,
            workflow_step=workflow_step,
            step_order=1,
            approver_user=cls.approver,
            acted_by=cls.approver,
            status="APPROVED",
            acted_at=timezone.now(),
        )

    def update_url(self, request_obj=None):
        request_obj = request_obj or self.request_obj
        return reverse("update_material_issue_note", args=[request_obj.id])

    def post_note(self, user, note, request_obj=None):
        self.client.force_login(user)
        return self.client.post(
            self.update_url(request_obj),
            {"material_issue_note": note},
            HTTP_HOST="127.0.0.1",
        )

    def test_stock_manager_can_add_and_edit_note_with_audit_log(self):
        self.request_obj.material_issue_note = ""
        self.request_obj.save(update_fields=["material_issue_note"])

        response = self.post_note(self.stock_manager, "SN-100")
        self.assertRedirects(
            response,
            reverse("request_detail", args=[self.request_obj.id]),
            fetch_redirect_response=False,
        )
        self.request_obj.refresh_from_db()
        self.assertEqual(self.request_obj.material_issue_note, "SN-100")

        self.post_note(self.stock_manager, "SN-101; installed")
        self.request_obj.refresh_from_db()
        self.assertEqual(self.request_obj.material_issue_note, "SN-101; installed")
        self.assertEqual(
            RequestAuditLog.objects.filter(
                request=self.request_obj,
                action="MATERIAL_ISSUE_NOTE_UPDATED",
                performed_by=self.stock_manager,
            ).count(),
            2,
        )

    def test_superuser_can_edit_note(self):
        self.post_note(self.superuser, "Updated by admin")
        self.request_obj.refresh_from_db()
        self.assertEqual(self.request_obj.material_issue_note, "Updated by admin")

    def test_stock_manager_sees_edit_form_and_french_label(self):
        self.client.force_login(self.stock_manager)
        response = self.client.get(
            reverse("request_detail", args=[self.request_obj.id]),
            HTTP_ACCEPT_LANGUAGE="fr",
            HTTP_HOST="127.0.0.1",
        )
        self.assertContains(response, "Note de sortie matériel")
        self.assertContains(response, "Enregistrer la note")
        self.assertContains(response, "SN-OLD")

    def test_requester_can_view_note_but_not_edit(self):
        self.client.force_login(self.requester)
        response = self.client.get(
            reverse("request_detail", args=[self.request_obj.id]),
            HTTP_HOST="127.0.0.1",
        )
        self.assertContains(response, "SN-OLD")
        self.assertNotContains(response, "Save Note")
        self.assertEqual(self.post_note(self.requester, "Forbidden").status_code, 403)

    def test_approver_can_view_note_but_not_edit(self):
        self.client.force_login(self.approver)
        response = self.client.get(
            reverse("request_detail", args=[self.request_obj.id]),
            HTTP_HOST="127.0.0.1",
        )
        self.assertContains(response, "SN-OLD")
        self.assertNotContains(response, "Save Note")
        self.assertEqual(self.post_note(self.approver, "Forbidden").status_code, 403)

    def test_non_stock_user_cannot_post_update(self):
        response = self.post_note(self.regular_user, "Forbidden")
        self.assertEqual(response.status_code, 403)
        self.request_obj.refresh_from_db()
        self.assertEqual(self.request_obj.material_issue_note, "SN-OLD")

    def test_non_approved_request_cannot_update_note(self):
        self.request_obj.status = "PENDING"
        self.request_obj.save(update_fields=["status"])
        response = self.post_note(self.stock_manager, "Forbidden")
        self.assertEqual(response.status_code, 302)
        self.request_obj.refresh_from_db()
        self.assertEqual(self.request_obj.material_issue_note, "SN-OLD")

    def test_non_material_request_cannot_update_note(self):
        request_obj = Request.objects.create(
            request_number="NOTE-002",
            request_type=self.non_material_type,
            submitted_by=self.requester,
            department=self.department,
            description="Service only",
            status="APPROVED",
        )
        response = self.post_note(self.stock_manager, "Forbidden", request_obj)
        self.assertEqual(response.status_code, 302)
        request_obj.refresh_from_db()
        self.assertEqual(request_obj.material_issue_note, "")

    def test_update_endpoint_is_post_only(self):
        self.client.force_login(self.stock_manager)
        response = self.client.get(self.update_url(), HTTP_HOST="127.0.0.1")
        self.assertEqual(response.status_code, 405)

    def test_note_appears_in_single_and_bulk_material_slips(self):
        self.client.force_login(self.stock_manager)
        single_response = self.client.get(
            reverse("approved_document", args=[self.request_obj.id]),
            HTTP_HOST="127.0.0.1",
        )
        self.assertContains(single_response, "Material Issue Note")
        self.assertContains(single_response, "SN-OLD")

        bulk_response = self.client.post(
            reverse("bulk_print_material_documents"),
            {"selected_requests": [self.request_obj.id]},
            HTTP_HOST="127.0.0.1",
        )
        self.assertContains(bulk_response, "Material Issue Note")
        self.assertContains(bulk_response, "SN-OLD")

    def test_note_appears_in_material_report_and_excel_export(self):
        self.client.force_login(self.stock_manager)
        report_response = self.client.get(
            reverse("material_reports"),
            HTTP_HOST="127.0.0.1",
        )
        self.assertContains(report_response, "SN-OLD")

        excel_response = self.client.get(
            reverse("export_material_report_excel"),
            HTTP_HOST="127.0.0.1",
        )
        sheet = load_workbook(BytesIO(excel_response.content)).active
        self.assertEqual(sheet["M4"].value, "Material Issue Note")
        self.assertEqual(sheet["M5"].value, "SN-OLD")


class ReturnedMaterialRequestEditTests(TestCase):
    def setUp(self):
        self.department = Department.objects.create(
            name="Operations",
            code="OPS",
        )

        User = get_user_model()
        self.requester = User.objects.create_user(
            username="returned_requester",
            password="pass12345",
            email="returned-requester@example.com",
            full_name="Returned Requester",
            department=self.department,
        )
        self.approver = User.objects.create_user(
            username="returned_approver",
            password="pass12345",
            email="returned-approver@example.com",
            full_name="Returned Approver",
        )

        self.request_type = RequestType.objects.create(
            name="Material Request",
            code="RETURNED-MATERIAL",
            is_active=True,
            requires_materials=True,
        )
        self.workflow = ApprovalWorkflow.objects.create(
            name="Returned material approval",
            request_type=self.request_type,
            department=self.department,
            is_active=True,
        )
        self.workflow_step = ApprovalWorkflowStep.objects.create(
            workflow=self.workflow,
            step_order=1,
            approver_user=self.approver,
        )

        self.category = MaterialCategory.objects.create(
            name="Networking",
            code="NETWORKING",
        )
        self.existing_material = Material.objects.create(
            category=self.category,
            name="Ethernet Cable",
            code="ETH-CABLE",
            unit="roll",
            stock_quantity=10,
        )
        self.new_material = Material.objects.create(
            category=self.category,
            name="Network Switch",
            code="NET-SWITCH",
            unit="pcs",
            stock_quantity=8,
        )

        self.request_obj = Request.objects.create(
            request_number="REQ-RETURNED-MATERIAL",
            request_type=self.request_type,
            submitted_by=self.requester,
            department=self.department,
            description="Network installation",
            date_needed=timezone.localdate(),
            status="RETURNED",
        )
        self.existing_item = RequestMaterialItem.objects.create(
            request=self.request_obj,
            material=self.existing_material,
            quantity=2,
        )
        RequestApproval.objects.create(
            request=self.request_obj,
            workflow_step=self.workflow_step,
            step_order=1,
            approver_user=self.approver,
            status="RETURNED",
        )

        self.client.force_login(self.requester)

    def edit_url(self):
        return reverse("edit_request", args=[self.request_obj.id])

    def request_data(self):
        return {
            "request_type": self.request_type.id,
            "description": "Corrected network installation",
            "date_needed": timezone.localdate().isoformat(),
        }

    def test_edit_page_uses_shared_searchable_material_picker(self):
        response = self.client.get(
            self.edit_url(),
            HTTP_HOST="127.0.0.1",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="material-category-filter"')
        self.assertContains(response, 'class="material-search-field"')
        self.assertContains(response, 'id="add-material-btn"')
        self.assertContains(response, 'id="empty-material-template"')
        self.assertContains(response, 'name="material_items-TOTAL_FORMS"')
        self.assertContains(response, "Ethernet Cable")
        self.assertContains(response, "ETH-CABLE")
        self.assertContains(response, "Networking")
        self.assertContains(response, "Stock: 10.00 roll")

    def test_create_and_edit_pages_use_safe_material_formset_controls(self):
        for url in (reverse("create_request"), self.edit_url()):
            with self.subTest(url=url):
                response = self.client.get(url, HTTP_HOST="127.0.0.1")
                html = response.content.decode()

                self.assertEqual(response.status_code, 200)
                self.assertIn(
                    '<button type="button" class="btn btn-secondary" id="add-material-btn">',
                    html,
                )
                self.assertIn(
                    '<button type="button" class="remove-material-btn">',
                    html,
                )
                self.assertIn('const initialForms = document.getElementById(', html)
                self.assertIn("deleteInput.checked = true;", html)
                self.assertIn("row.remove();", html)
                self.assertIn("totalForms.value = formIndex + 1;", html)
                self.assertNotIn("addMaterialRow();", html)

    def test_create_accepts_form_index_gap_left_by_removed_unsaved_rows(self):
        data = self.request_data()
        data.update(
            {
                "material_items-TOTAL_FORMS": "3",
                "material_items-INITIAL_FORMS": "0",
                "material_items-MIN_NUM_FORMS": "0",
                "material_items-MAX_NUM_FORMS": "1000",
                "material_items-2-material": self.new_material.id,
                "material_items-2-quantity": "3",
            }
        )

        response = self.client.post(
            reverse("create_request"),
            data=data,
            HTTP_HOST="127.0.0.1",
        )

        self.assertRedirects(response, reverse("dashboard"))
        created_request = Request.objects.exclude(id=self.request_obj.id).get()
        self.assertEqual(created_request.material_items.count(), 1)
        self.assertEqual(created_request.material_items.get().material, self.new_material)
        self.assertEqual(created_request.material_items.get().quantity, 3)

    def test_resubmit_can_remove_existing_item_and_add_new_item(self):
        data = self.request_data()
        data.update(
            {
                "material_items-TOTAL_FORMS": "3",
                "material_items-INITIAL_FORMS": "1",
                "material_items-MIN_NUM_FORMS": "0",
                "material_items-MAX_NUM_FORMS": "1000",
                "material_items-0-id": self.existing_item.id,
                "material_items-0-material": self.existing_material.id,
                "material_items-0-quantity": "2",
                "material_items-0-DELETE": "on",
                "material_items-2-material": self.new_material.id,
                "material_items-2-quantity": "3",
            }
        )

        response = self.client.post(
            self.edit_url(),
            data=data,
            HTTP_HOST="127.0.0.1",
        )

        self.assertRedirects(
            response,
            reverse("request_detail", args=[self.request_obj.id]),
            fetch_redirect_response=False,
        )

        self.request_obj.refresh_from_db()
        self.existing_material.refresh_from_db()
        self.new_material.refresh_from_db()

        self.assertEqual(self.request_obj.status, "IN_REVIEW")
        self.assertEqual(self.request_obj.material_items.count(), 1)
        self.assertEqual(
            self.request_obj.material_items.get().material,
            self.new_material,
        )
        self.assertEqual(self.request_obj.material_items.get().quantity, 3)
        self.assertEqual(self.request_obj.approvals.count(), 1)
        self.assertEqual(self.request_obj.approvals.get().status, "PENDING")
        self.assertEqual(self.existing_material.stock_quantity, 10)
        self.assertEqual(self.new_material.stock_quantity, 8)
        self.assertFalse(self.request_obj.stock_deducted)

    def test_resubmit_rejects_quantity_above_available_stock(self):
        data = self.request_data()
        data.update(
            {
                "material_items-TOTAL_FORMS": "2",
                "material_items-INITIAL_FORMS": "1",
                "material_items-MIN_NUM_FORMS": "0",
                "material_items-MAX_NUM_FORMS": "1000",
                "material_items-0-id": self.existing_item.id,
                "material_items-0-material": self.existing_material.id,
                "material_items-0-quantity": "11",
            }
        )

        response = self.client.post(
            self.edit_url(),
            data=data,
            HTTP_HOST="127.0.0.1",
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Only 10.00 roll available in stock.")
        self.assertContains(
            response,
            f'href="{reverse("request_detail", args=[self.request_obj.id])}"',
        )

        self.request_obj.refresh_from_db()
        self.existing_item.refresh_from_db()
        self.assertEqual(self.request_obj.status, "RETURNED")
        self.assertEqual(self.existing_item.quantity, 2)
        self.assertEqual(self.request_obj.approvals.get().status, "RETURNED")
